import os
import re
import math
import time
import subprocess
import argparse
import sys
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    HAS_GUI = True
except ImportError:
    HAS_GUI = False

from openpyxl import Workbook, load_workbook
import undetected_chromedriver as uc
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException

import re
import subprocess

def _parse_major(version_text: str):
    # Accepts: "Google Chrome 144.0.7559.110" or "Chrome 144.0...."
    m = re.search(r"(\d+)\.", version_text or "")
    return int(m.group(1)) if m else None


def get_chrome_major_version():
    """
    Robust Windows + Linux Chrome major version detection.
    """
    # --- 1) Try Registry (Windows Only)
    if os.name == 'nt':
        reg_keys = [
            r"HKCU\Software\Google\Chrome\BLBeacon",
            r"HKLM\Software\Google\Chrome\BLBeacon",
            r"HKLM\Software\WOW6432Node\Google\Update\Clients\{8A69D345-D564-463c-AFF1-A69D9E530F96}",
        ]
        for key in reg_keys:
            try:
                reg_out = subprocess.check_output(["reg", "query", key, "/v", "version"], text=True, stderr=subprocess.STDOUT)
                m = re.search(r"REG_SZ\s+([0-9.]+)", reg_out) or re.search(r"version\s+REG_SZ\s+([0-9.]+)", reg_out, re.IGNORECASE)
                if m:
                    major = _parse_major(m.group(1))
                    if major:
                        print(f"DEBUG: Found Chrome {major} via Registry.")
                        return major
            except Exception:
                pass

    # --- 2) Try CLI commands (Windows & Linux)
    commands = [
        ["google-chrome", "--version"],
        ["chrome", "--version"],
        ["google-chrome-stable", "--version"],
    ]
    for cmd in commands:
        try:
            ver_out = subprocess.check_output(cmd, text=True, stderr=subprocess.STDOUT).strip()
            major = _parse_major(ver_out)
            if major:
                print(f"DEBUG: Found Chrome {major} via CLI ({' '.join(cmd)})")
                return major
        except Exception:
            pass

    # --- 3) Try absolute paths (Windows)
    if os.name == 'nt':
        local = os.environ.get("LOCALAPPDATA", "")
        candidates = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            os.path.join(local, r"Google\Chrome\Application\chrome.exe"),
        ]
        for chrome_path in candidates:
            if os.path.exists(chrome_path):
                try:
                    ver_out = subprocess.check_output([chrome_path, "--version"], text=True).strip()
                    major = _parse_major(ver_out)
                    if major:
                        print(f"DEBUG: Found Chrome {major} via Path.")
                        return major
                except Exception:
                    pass
    return None


PER_PAGE = 24
from datetime import datetime

# Generate a timestamped filename for unique exports
_now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
OUTPUT_FILENAME = f"Echovita_Results_{_now}.xlsx"
CAPTCHA_TEXT = (
    "This website uses a security service to protect against malicious bots. "
    "This page is displayed while the website verifies you are not a bot."
)


# ---------------------- TKINTER HELPERS ---------------------- #

def make_hidden_root():
    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    return root


def select_input_file():
    # Directly return the input file to skip human interaction/dialogs
    # Even if HAS_GUI is true, we skip it as requested.
    filename = "Echovita_Input.txt"
    if os.path.exists(filename):
        return filename
    
    # Fallback if the file doesn't exist (though usually it should)
    if os.environ.get("GITHUB_ACTIONS") or not HAS_GUI:
        return filename

    # Optional: only show dialog if file NOT found
    root = make_hidden_root()
    file_path = filedialog.askopenfilename(
        title="Select Echovita Input TXT File",
        filetypes=[("Text Files", "*.txt"), ("All Files", "*.*")]
    )
    root.destroy()
    return file_path or filename


def show_info(title, message):
    # Strictly use print instead of message boxes for automation
    print(f"\n[INFO] {title}: {message}\n")


def show_error(title, message):
    # Strictly use print instead of message boxes for automation
    print(f"\n[ERROR] {title}: {message}\n")


# ---------------------- EXCEL HELPERS ---------------------- #

def get_output_file_path(input_file_path):
    input_dir = os.path.dirname(os.path.abspath(input_file_path))
    return os.path.join(input_dir, OUTPUT_FILENAME)


def create_or_load_workbook(output_path):
    if os.path.exists(output_path):
        try:
            return load_workbook(output_path)
        except Exception:
            raise Exception(
                f"Output file exists but is not a valid Excel workbook:\n{output_path}\n"
                f"Delete it and run again."
            )

    wb = Workbook()
    ws = wb.active
    ws.title = "temp"
    wb.save(output_path)
    return wb


def ensure_sheet(wb, sheet_name, current_url, output_path):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_found = True
    else:
        ws = wb.create_sheet(title=sheet_name)
        sheet_found = False

        headers = [
            "S.no.", "Input URL", "Name", "DOD", "Age", "Record URL",
            "TotalCount", "LastSerial", "LastPage", "LastRecordIndex",
            "Total Pages", "Sheet URL"
        ]
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        ws.cell(row=2, column=12, value=current_url)

        if "temp" in wb.sheetnames and len(wb.sheetnames) > 1:
            temp_ws = wb["temp"]
            wb.remove(temp_ws)

        wb.save(output_path)

    return ws, sheet_found


def read_int_cell(ws, row, col, default=0):
    value = ws.cell(row=row, column=col).value
    if value is None or value == "":
        return default
    if str(value).strip() == "Page not found":
        return default
    try:
        return int(float(str(value).strip()))
    except Exception:
        return default


def read_text_cell(ws, row, col, default=""):
    value = ws.cell(row=row, column=col).value
    if value is None:
        return default
    return str(value)


def save_workbook_safe(wb, output_path):
    wb.save(output_path)


# ---------------------- SELENIUM / PAGE HELPERS ---------------------- #

import os

def create_driver():
    is_ci = bool(os.environ.get("GITHUB_ACTIONS"))
    chrome_major = get_chrome_major_version()
    
    options = uc.ChromeOptions()
    options.add_argument("--start-maximized")
    options.add_argument("--log-level=3")
    options.add_argument("--disable-logging")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--disable-background-timer-throttling")
    options.add_argument("--disable-backgrounding-occluded-windows")
    options.add_argument("--disable-renderer-backgrounding")
    options.add_argument("--disable-features=CalculateNativeWinOcclusion")

    if is_ci:
        options.add_argument("--headless=new")
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--window-size=1920,1080")

    driver = uc.Chrome(options=options, version_main=chrome_major)
    driver.set_page_load_timeout(60)
    return driver

def get_body_text(driver):
    try:
        return driver.find_element("tag name", "body").text or ""
    except Exception:
        return ""


def handle_captcha_if_present(driver):
    is_ci = bool(os.environ.get("GITHUB_ACTIONS"))
    while True:
        body_text = get_body_text(driver)
        if CAPTCHA_TEXT in body_text:
            if is_ci:
                print("CAPTCHA DETECTED in CI environment! Skipping or failing...")
                raise Exception("CAPTCHA detected in headless environment (CI).")
            show_info("Captcha Detected", "Solve captcha and click OK")
            time.sleep(2)
            continue
        break


def open_url(driver, url, expect_text=None, timeout=60):
    driver.get(url)
    time.sleep(3)
    handle_captcha_if_present(driver)

    if expect_text:
        def condition(d):
            body = get_body_text(d).lower()
            return expect_text.lower() in body or "page not found" in body

        WebDriverWait(driver, timeout).until(condition)
        handle_captcha_if_present(driver)

def page_contains_text(driver, text):
    return text.lower() in get_body_text(driver).lower()


def build_page_url(current_url, current_page):
    if re.search(r'([?&])page=\d+', current_url, flags=re.I):
        return re.sub(r'([?&])page=\d+', rf'\1page={current_page}', current_url, flags=re.I)

    if "?" in current_url:
        return f"{current_url}&page={current_page}"
    return f"{current_url}?page={current_page}"


# ---------------------- JAVASCRIPT SCRAPERS ---------------------- #

JS_TOTAL_COUNT = r"""
function scrapeObitCount() {
  const root = document.querySelector('#desktopObitSearch');
  if (!root) return '';
  const el = root.querySelector('div.d-flex.flex-column > span.eh-font-weight400')
           || root.querySelector('div.d-flex.flex-column > span');
  const txt = (el?.textContent || '').replace(/\s+/g, ' ').trim();
  return txt;
}
return scrapeObitCount();
"""

JS_RECORD_LINKS = r"""
function getRecordLinks() {
  var origin = location.origin;
  var nodes = document.querySelectorAll('.obit-list-wrapper a.text-name-obit-in-list[href]');
  if (!nodes || !nodes.length) return '';

  var seen = Object.create(null), out = [];
  for (var i = 0; i < nodes.length; i++) {
    var href = (nodes[i].getAttribute('href') || '').trim();
    if (!href) continue;
    if (href.charAt(0) === '/') href = origin + href;
    if (!seen[href]) { seen[href] = true; out.push(href); }
  }

  return out.join('|') + '';
}
return getRecordLinks();
"""

JS_NAME = r"""
function getName() {
  try {
    var el =
      document.querySelector('.obit-main-info-wrapper-min-height p.my-auto.h1.text-white.font-weight-bolder')
      || document.querySelector('h1.text-center.text-lg-left.font-weight-bolder.mb-2')
      || document.querySelector('h1, .display-4.font-weight-bolder');

    var txt = (el && el.textContent || '').replace(/\s+/g, ' ').trim();
    return txt + '';
  } catch (e) {
    return '';
  }
}
return getName();
"""

JS_DOD = r"""
function getDOD() {
  try {
    var p = document.querySelector('.obit-main-info-wrapper-min-height p.mt-2.mb-1.text-white.font-weight-bold');
    var raw = (p && p.textContent || '').replace(/\s+/g, ' ').trim();
    if (!raw) return '';

    var dates = raw.match(/([A-Za-z]+ \d{1,2}, \d{4})/g);
    if (dates && dates.length) {
      return (dates[dates.length - 1] + '');
    }
    return '';
  } catch (e) {
    return '';
  }
}
return getDOD();
"""

JS_AGE = r"""
function getAge() {
  try {
    var p = document.querySelector('.obit-main-info-wrapper-min-height p.mt-2.mb-1.text-white.font-weight-bold');
    var raw = (p && p.textContent || '').replace(/\s+/g, ' ').trim();
    if (!raw) return '';

    var m = raw.match(/\((\d{1,3})\s*years?\s*old\)/i);
    if (m) return (m[1] + '');

    m = raw.match(/\b(\d{1,3})\s*years?\b/i);
    if (m) return (m[1] + '');

    return '';
  } catch (e) {
    return '';
  }
}
return getAge();
"""


def get_total_count(driver):
    label = driver.execute_script(JS_TOTAL_COUNT) or ""
    match = re.search(r'(\d[\d,]*)', str(label))
    if not match:
        return 0
    try:
        return int(match.group(1).replace(",", ""))
    except Exception:
        return 0


def get_record_links(driver):
    raw = driver.execute_script(JS_RECORD_LINKS) or ""
    raw = str(raw).strip()
    if not raw:
        return []
    return [x.strip() for x in raw.split("|") if x.strip()]


def get_name(driver):
    return (driver.execute_script(JS_NAME) or "").strip()


def get_dod(driver):
    return (driver.execute_script(JS_DOD) or "").strip()


def get_age(driver):
    return (driver.execute_script(JS_AGE) or "").strip()


# ---------------------- MAIN LOGIC ---------------------- #

def run():
    driver = None
    wb = None
    output_path = None

    try:
        input_file_path = select_input_file()
        if not input_file_path:
            show_info("Cancelled", "No input file selected.")
            return

        output_path = get_output_file_path(input_file_path)
        wb = create_or_load_workbook(output_path)

        with open(input_file_path, "r", encoding="utf-8") as f:
            url_list = [line.strip() for line in f.read().splitlines() if line.strip()]

        if not url_list:
            show_info("No URLs", "Input file is empty.")
            return

        print(f"Total input URLs found: {len(url_list)}")

        serial = 1

        for url_index, current_url in enumerate(url_list, start=1):
            print(f"\nProcessing Url {url_index}: {current_url}")

            sheet_name = str(url_index)
            ws, sheet_found = ensure_sheet(wb, sheet_name, current_url, output_path)

            if not sheet_found:
                next_write_row = 2
                last_serial_number = 0
            else:
                last_serial_stored_cell = ws.cell(row=2, column=8).value
                if last_serial_stored_cell in (None, ""):
                    last_serial_number = 0
                else:
                    try:
                        last_serial_number = int(float(str(last_serial_stored_cell).strip()))
                    except Exception:
                        last_serial_number = 0

                next_write_row = last_serial_number + 2
                if next_write_row < 2:
                    next_write_row = 2

            sheet_url_cell = ws.cell(row=2, column=12).value
            if sheet_url_cell in (None, ""):
                ws.cell(row=2, column=12, value=current_url)
                save_workbook_safe(wb, output_path)
                sheet_url = current_url
            else:
                sheet_url = str(sheet_url_cell)

            total_count_for_url = read_int_cell(ws, 2, 7, 0)
            last_serial_stored = read_int_cell(ws, 2, 8, 0)
            last_page_stored = read_int_cell(ws, 2, 9, 0)
            last_record_index_stored = read_int_cell(ws, 2, 10, 0)

            total_count_raw = read_text_cell(ws, 2, 7, "")
            if total_count_raw == "Page not found":
                total_count_for_url = 0

            if last_serial_number < last_serial_stored:
                last_serial_number = last_serial_stored

            if driver is None:
                driver = create_driver()

            try:
                open_url(driver, current_url, expect_text="Receive obituaries", timeout=120)
            except TimeoutException:
                if page_contains_text(driver, "Page not found"):
                    print(f"Url {url_index}: Page not found")
                    ws.cell(row=2, column=7, value="Page not found")
                    save_workbook_safe(wb, output_path)
                    continue
                raise

            if page_contains_text(driver, "Page not found"):
                print(f"Url {url_index}: Page not found")
                ws.cell(row=2, column=7, value="Page not found")
                save_workbook_safe(wb, output_path)
                continue

            if total_count_for_url == 0:
                total_count_for_url = get_total_count(driver)
                ws.cell(row=2, column=7, value=total_count_for_url)
                save_workbook_safe(wb, output_path)
                print(f"Found {total_count_for_url} records in Url {url_index}")
                time.sleep(2)

            if total_count_for_url > 0:
                print(f"Found {total_count_for_url} records in Url {url_index}")

            skip_this_url = False
            if total_count_for_url > 0 and last_serial_number >= total_count_for_url:
                ws.cell(row=2, column=8, value=last_serial_number)
                save_workbook_safe(wb, output_path)
                print(f"Url {url_index} already completed. Skipping.")
                skip_this_url = True

            if skip_this_url:
                continue

            start_page = last_page_stored if last_page_stored > 0 else 1
            start_record_index = last_record_index_stored + 1 if last_record_index_stored > 0 else 1

            current_page = start_page
            record_serial = last_serial_number + 1

            total_pages = 0 if total_count_for_url <= 0 else math.ceil(total_count_for_url / PER_PAGE)
            ws.cell(row=2, column=11, value=total_pages)
            save_workbook_safe(wb, output_path)

            if total_pages <= 0:
                continue

            for current_page in range(1, total_pages + 1):
                if current_page < start_page:
                    continue

                page_url = build_page_url(current_url, current_page)
                open_url(driver, page_url, expect_text="Receive obituaries", timeout=120)
                print(f"Opened page {current_page} for Url {url_index}")

                records_list = get_record_links(driver)
                print(f"Found {len(records_list)} record links on page {current_page} for Url {url_index}")
                record_index = 1

                for record_url in records_list:
                    if current_page == start_page and record_index < start_record_index:
                        record_index += 1
                        continue

                    open_url(driver, record_url, expect_text="Obituary", timeout=120)

                    name_text = get_name(driver)
                    dod_text = get_dod(driver)
                    age_text = get_age(driver)

                    ws.cell(row=next_write_row, column=1, value=record_serial)
                    ws.cell(row=next_write_row, column=2, value=sheet_url)
                    ws.cell(row=next_write_row, column=3, value=name_text)
                    ws.cell(row=next_write_row, column=4, value=dod_text)
                    ws.cell(row=next_write_row, column=5, value=age_text)
                    ws.cell(row=next_write_row, column=6, value=record_url)

                    ws.cell(row=2, column=8, value=record_serial)
                    ws.cell(row=2, column=9, value=current_page)
                    ws.cell(row=2, column=10, value=record_index)

                    save_workbook_safe(wb, output_path)
                    print(f"Saved Record {record_serial} with name {name_text}")

                    record_serial += 1
                    next_write_row += 1
                    record_index += 1

            serial += 1

        save_workbook_safe(wb, output_path)
        show_info("All URLs Done", f"Output saved in excel:\n{output_path}")

    except Exception as e:
        err_text = (
            f"An error occurred.\n\n"
            f"Error: {str(e)}\n\n"
            f"The tool will now close safely."
        )
        print(f"ERROR: {str(e)}")
        show_error("Tool Error", err_text)

    finally:
        try:
            if wb and output_path:
                wb.save(output_path)
        except Exception:
            pass

        try:
            if driver:
                driver.quit()
        except Exception:
            pass


if __name__ == "__main__":
    run()